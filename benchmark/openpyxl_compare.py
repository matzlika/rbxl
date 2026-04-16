#!/usr/bin/env python3

import json
import os
import resource
import sys
import tempfile
import time

from openpyxl import Workbook, load_workbook


ROWS = int(os.environ.get("RBXL_BENCH_ROWS", "5000"))
COLS = int(os.environ.get("RBXL_BENCH_COLS", "10"))
WARMUP = int(os.environ.get("RBXL_BENCH_WARMUP", "1"))
ITERATIONS = int(os.environ.get("RBXL_BENCH_ITERATIONS", "5"))
READ_PATH = os.environ.get("RBXL_BENCH_READ_PATH")


def rss_kb() -> int:
    return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss


def build_dataset(rows: int, cols: int):
    header = [f"col_{i + 1}" for i in range(cols)]
    body = []
    for row in range(rows):
      values = []
      for col in range(cols):
          mod = col % 4
          if mod == 0:
              values.append(row)
          elif mod == 1:
              values.append(f"row-{row}-col-{col}")
          elif mod == 2:
              values.append((row + col) % 2 == 1)
          else:
              values.append(((row * 100) + col) / 10.0)
      body.append(values)
    return header, body


def measure(label, func):
    for _ in range(WARMUP):
        func()

    samples = []
    rss_deltas = []
    result = None
    for _ in range(ITERATIONS):
        before = rss_kb()
        started = time.perf_counter()
        result = func()
        elapsed = time.perf_counter() - started
        samples.append(elapsed)
        rss_deltas.append(max(0, rss_kb() - before))

    mean = sum(samples) / len(samples)
    variance = sum((sample - mean) ** 2 for sample in samples) / len(samples)
    return {
        "label": label,
        "real": mean,
        "real_min": min(samples),
        "real_stddev": variance ** 0.5,
        "rss_delta_kb": max(rss_deltas),
        "iterations": ITERATIONS,
        "result": result,
    }


def write_with_openpyxl(path, header, body):
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Bench")
    sheet.append(header)
    for row in body:
        sheet.append(row)
    workbook.save(path)


def read_with_openpyxl(path):
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook["Bench"]
    count = 0
    for row in sheet.rows:
        count += len(row)
    workbook.close()
    return count


def read_values_with_openpyxl(path):
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook["Bench"]
    count = 0
    for row in sheet.iter_rows(values_only=True):
        count += len(row)
    workbook.close()
    return count


def main():
    header, body = build_dataset(ROWS, COLS)
    with tempfile.TemporaryDirectory(prefix="rbxl-openpyxl-") as directory:
        path = os.path.join(directory, "openpyxl.xlsx")
        results = []

        write_result = measure("openpyxl write", lambda: write_with_openpyxl(path, header, body))
        write_result["size"] = os.path.getsize(path)
        results.append(write_result)
        read_path = READ_PATH or path
        results.append(measure("openpyxl read", lambda: read_with_openpyxl(read_path)))
        results.append(measure("openpyxl read values", lambda: read_values_with_openpyxl(read_path)))

        print(json.dumps(results))


if __name__ == "__main__":
    sys.exit(main())
